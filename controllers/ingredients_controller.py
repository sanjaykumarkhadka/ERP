import pandas as pd
import os
from flask import Blueprint, jsonify, render_template, request, redirect, url_for, flash
from sqlalchemy.sql import text
from sqlalchemy import asc, desc
from werkzeug.utils import secure_filename
from datetime import datetime, date
import pytz

ingredients_bp = Blueprint('ingredients', __name__, template_folder='templates')

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_rm_type_id():
    """Helper function to get RM item type ID"""
    from models.item_type import ItemType
    rm_type = ItemType.query.filter_by(type_name='RM').first()
    return rm_type.id if rm_type else None

@ingredients_bp.route('/ingredients_list', methods=['GET'])
def ingredients_list():
    from app import db
    from models.item_master import ItemMaster
    from models.raw_material_stocktake import RawMaterialStocktake
    from models.category import Category
    from models.department import Department
    from models.uom import UOM
    from models.item_type import ItemType

    # Get search parameters
    search_item_code = request.args.get('item_code', '').strip()
    search_description = request.args.get('description', '').strip()
    search_category = request.args.get('category', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 25  # Number of items per page
    
    # Get RM type ID
    rm_type_id = get_rm_type_id()
    if not rm_type_id:
        flash("Raw Material item type not found in system!", "danger")
        return render_template('ingredients/list.html', ingredients=[], categories=[], current_page="ingredients")

    # Build query for ingredients (Raw Materials only)
    ingredients_query = ItemMaster.query.filter(ItemMaster.item_type_id == rm_type_id)
    
    if search_item_code:
        ingredients_query = ingredients_query.filter(ItemMaster.item_code.ilike(f"%{search_item_code}%"))
    if search_description:
        ingredients_query = ingredients_query.filter(ItemMaster.description.ilike(f"%{search_description}%"))
    
    # Add category filter if specified
    if search_category:
        try:
            category_id = int(search_category)
            ingredients_query = ingredients_query.filter(ItemMaster.category_id == category_id)
        except (ValueError, TypeError):
            pass

    # Execute paginated query
    ingredients = ingredients_query.paginate(
        page=page, per_page=per_page, error_out=False
    )

    # Get categories for dropdown
    categories = Category.query.all()

    return render_template('ingredients/list.html',
                           ingredients=ingredients,
                           categories=categories,
                           current_page="ingredients",
                           search_item_code=search_item_code,
                           search_description=search_description,
                           search_category=search_category)

@ingredients_bp.route('/ingredients_create', methods=['GET', 'POST'])
def ingredients_create():
    from app import db
    from models.item_master import ItemMaster
    from models.category import Category
    from models.department import Department
    from models.uom import UOM
    from models.allergen import Allergen
    from models.raw_material_stocktake import RawMaterialStocktake
    from datetime import datetime, date

    if request.method == 'POST':
        try:
            # Get form data
            week_commencing_str = request.form.get('week_commencing', '').strip()
            stocktake_type = request.form.get('stocktake_type', '').strip()
            user = request.form.get('user', '').strip()
            item_code = request.form['item_code'].strip()
            category_id = request.form.get('category_id') or None
            current_stock = float(request.form.get('current_stock') or 0.0)
            order_quantity = float(request.form.get('order_quantity') or 0.0)
            price_uom = float(request.form.get('price_uom') or 0.0)
            notes = request.form.get('notes', '').strip()

            # Validate required fields
            if not week_commencing_str:
                flash("Week commencing is required!", "danger")
                return redirect(request.url)
            
            if not stocktake_type:
                flash("Stocktake type is required!", "danger")
                return redirect(request.url)
                
            if not user:
                flash("User is required!", "danger")
                return redirect(request.url)

            # Parse week commencing date
            try:
                week_commencing = datetime.strptime(week_commencing_str, '%Y-%m-%d').date()
            except ValueError:
                flash("Invalid week commencing date format!", "danger")
                return redirect(request.url)

            # Check if item exists in Item Master and is a raw material
            rm_type_id = get_rm_type_id()
            if not rm_type_id:
                flash("RM item type not found in system!", "danger")
                return redirect(request.url)
                
            existing_item = ItemMaster.query.filter_by(item_code=item_code, item_type_id=rm_type_id).first()
            if not existing_item:
                flash(f"Raw material '{item_code}' not found in Item Master!", "danger")
                return redirect(request.url)

            # Calculate stock value
            stock_value = current_stock * price_uom

            # Create new stocktake record
            new_stocktake = RawMaterialStocktake(
                week_commencing=week_commencing,
                stocktake_type=stocktake_type,
                user=user,
                item_code=item_code,
                category_id=int(category_id) if category_id else None,
                current_stock=current_stock,
                order_quantity=order_quantity,
                price_uom=price_uom,
                stock_value=stock_value,
                notes=notes
            )
            
            db.session.add(new_stocktake)
            db.session.commit()
            
            flash(f"Stocktake record created successfully for {item_code}: {current_stock} units, {order_quantity} to order, worth ${stock_value:.2f}", "success")
            return redirect(url_for('ingredients.ingredients_list'))

        except ValueError as e:
            flash(f"Invalid number format: {str(e)}", "danger")
            return redirect(request.url)
        except Exception as e:
            db.session.rollback()
            flash(f"Error adding stock record: {str(e)}", "danger")
            return redirect(request.url)

    # Get existing raw materials from item_master table
    rm_type_id = get_rm_type_id()
    if rm_type_id:
        existing_items_query = ItemMaster.query.filter_by(item_type_id=rm_type_id).all()
    else:
        existing_items_query = []
    
    # Convert to dictionaries for JSON serialization
    existing_items = []
    for item in existing_items_query:
        existing_items.append({
            'id': item.id,
            'item_code': item.item_code,
            'description': item.description or '',
            'item_type': item.item_type.type_name if item.item_type else None,
            'category_id': item.category_id,
            'department_id': item.department_id,
            'uom_id': item.uom_id,
            'min_level': item.min_level or 0,
            'max_level': item.max_level or 0,
            'price_per_kg': item.price_per_kg or 0,
            'is_active': item.is_active
        })
    
    # Get departments, UOMs, and categories for dropdowns
    departments = Department.query.all()
    uoms = UOM.query.all()
    categories = Category.query.all()

    return render_template('ingredients/create.html',
                           existing_items=existing_items,
                           departments=departments,
                           uoms=uoms,
                           categories=categories,
                           current_page="ingredients")

@ingredients_bp.route('/ingredients_edit/<int:id>', methods=['GET', 'POST'])
def ingredients_edit(id):
    from app import db
    from models.item_master import ItemMaster, ItemAllergen
    from models.category import Category
    from models.department import Department
    from models.uom import UOM
    from models.allergen import Allergen

    rm_type_id = get_rm_type_id()
    if not rm_type_id:
        flash("RM item type not found in system!", "danger")
        return redirect(url_for('ingredients.ingredients_list'))

    ingredient = ItemMaster.query.filter_by(id=id, item_type_id=rm_type_id).first_or_404()

    if request.method == 'POST':
        try:
            item_code = request.form['item_code'].strip()
            description = request.form['description'].strip()
            category_id = request.form.get('category_id') or None
            department_id = request.form.get('department_id') or None
            uom_id = request.form.get('uom_id') or None
            
            min_level = float(request.form.get('min_level') or 0.0)
            max_level = float(request.form.get('max_level') or 0.0)
            price_per_kg = float(request.form.get('price_per_kg') or 0.0)
            is_active = 'is_active' in request.form

            # Check for duplicate item code (excluding current item)
            existing_item = ItemMaster.query.filter(
                ItemMaster.item_code == item_code,
                ItemMaster.id != id
            ).first()
            if existing_item:
                flash(f"Item code '{item_code}' already exists!", "danger")
                return redirect(request.url)

            # Update ingredient
            ingredient.item_code = item_code
            ingredient.description = description
            ingredient.category_id = int(category_id) if category_id else None
            ingredient.department_id = int(department_id) if department_id else None
            ingredient.uom_id = int(uom_id) if uom_id else None
            ingredient.min_level = min_level
            ingredient.max_level = max_level
            ingredient.price_per_kg = price_per_kg
            ingredient.is_active = is_active

            # Handle allergens
            # Remove existing allergen associations
            ItemAllergen.query.filter_by(item_id=ingredient.id).delete()
            
            # Add new allergen associations
            allergen_ids = request.form.getlist('allergen_ids')
            for allergen_id in allergen_ids:
                item_allergen = ItemAllergen(
                    item_id=ingredient.id,
                    allergen_id=int(allergen_id)
                )
                db.session.add(item_allergen)

            db.session.commit()
            flash("Ingredient updated successfully!", "success")
            return redirect(url_for('ingredients.ingredients_list'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error updating ingredient: {str(e)}", "danger")
            return redirect(request.url)

    # Get reference data for dropdowns
    categories = Category.query.all()
    departments = Department.query.all()
    uoms = UOM.query.all()
    allergens = Allergen.query.all()
    
    # Get current allergen associations
    current_allergen_ids = [ia.allergen_id for ia in ItemAllergen.query.filter_by(item_id=ingredient.id).all()]

    return render_template('ingredients/edit.html',
                           ingredient=ingredient,
                           categories=categories,
                           departments=departments,
                           uoms=uoms,
                           allergens=allergens,
                           current_allergen_ids=current_allergen_ids,
                           current_page="ingredients")

@ingredients_bp.route('/stocktake_edit/<int:id>', methods=['GET', 'POST'])
def stocktake_edit(id):
    from app import db
    from models.raw_material_stocktake import RawMaterialStocktake
    from models.item_master import ItemMaster
    from models.department import Department
    from models.uom import UOM
    from models.category import Category

    stocktake = RawMaterialStocktake.query.get_or_404(id)

    if request.method == 'POST':
        try:
            # Get form data
            week_commencing_str = request.form.get('week_commencing', '').strip()
            stocktake_type = request.form.get('stocktake_type', '').strip()
            user = request.form.get('user', '').strip()
            category_id = request.form.get('category_id') or None
            current_stock = float(request.form.get('current_stock') or 0.0)
            order_quantity = float(request.form.get('order_quantity') or 0.0)
            price_uom = float(request.form.get('price_uom') or 0.0)
            notes = request.form.get('notes', '').strip()

            # Validate required fields
            if not week_commencing_str or not stocktake_type or not user:
                flash("Week commencing, stocktake type, and user are required!", "danger")
                return redirect(request.url)

            # Parse week commencing date
            try:
                week_commencing = datetime.strptime(week_commencing_str, '%Y-%m-%d').date()
            except ValueError:
                flash("Invalid week commencing date format!", "danger")
                return redirect(request.url)

            # Update stocktake record
            stocktake.week_commencing = week_commencing
            stocktake.stocktake_type = stocktake_type
            stocktake.user = user
            stocktake.category_id = int(category_id) if category_id else None
            stocktake.current_stock = current_stock
            stocktake.order_quantity = order_quantity
            stocktake.price_uom = price_uom
            stocktake.stock_value = current_stock * price_uom
            stocktake.notes = notes
            
            db.session.commit()
            
            flash(f"Stocktake record updated successfully for {stocktake.item_code}: {current_stock} units, {order_quantity} to order, worth ${stocktake.stock_value:.2f}", "success")
            return redirect(url_for('ingredients.ingredients_list'))

        except ValueError as e:
            flash(f"Invalid number format: {str(e)}", "danger")
            return redirect(request.url)
        except Exception as e:
            db.session.rollback()
            flash(f"Error updating stocktake record: {str(e)}", "danger")
            return redirect(request.url)

    # Get departments, UOMs, and categories for dropdowns
    departments = Department.query.all()
    uoms = UOM.query.all()
    categories = Category.query.all()

    return render_template('ingredients/stocktake_edit.html',
                           stocktake=stocktake,
                           departments=departments,
                           uoms=uoms,
                           categories=categories,
                           current_page="ingredients")

@ingredients_bp.route('/stocktake_delete/<int:id>', methods=['POST'])
def stocktake_delete(id):
    from app import db
    from models.raw_material_stocktake import RawMaterialStocktake

    try:
        stocktake = RawMaterialStocktake.query.get_or_404(id)
        
        # Delete stocktake record
        db.session.delete(stocktake)
        db.session.commit()
        flash("Stocktake record deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting stocktake record: {str(e)}", "danger")

    return redirect(url_for('ingredients.ingredients_list'))

@ingredients_bp.route('/ingredients_delete/<int:id>', methods=['POST'])
def ingredients_delete(id):
    from app import db
    from models.item_master import ItemMaster

    try:
        rm_type_id = get_rm_type_id()
        if not rm_type_id:
            flash("RM item type not found in system!", "danger")
            return redirect(url_for('ingredients.ingredients_list'))

        ingredient = ItemMaster.query.filter_by(id=id, item_type_id=rm_type_id).first_or_404()
        
        db.session.delete(ingredient)
        db.session.commit()
        flash(f"Ingredient '{ingredient.item_code}' deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting ingredient: {str(e)}", "danger")
    
    return redirect(url_for('ingredients.ingredients_list'))

@ingredients_bp.route('/ingredients_upload', methods=['GET', 'POST'])
def ingredients_upload():
    from app import db
    from models.item_master import ItemMaster
    from models.raw_material_stocktake import RawMaterialStocktake
    from models.category import Category
    from models.department import Department
    from models.uom import UOM
    from datetime import datetime

    if request.method == 'POST':
        if 'file' not in request.files:
            flash("No file uploaded!", "danger")
            return render_template('ingredients/upload.html', current_page="ingredients")

        file = request.files['file']
        sheet_name = request.form.get('sheet_name', '').strip() or 'Stocktake'

        if file.filename == '':
            flash("No file selected!", "danger")
            return render_template('ingredients/upload.html', current_page="ingredients")

        if not file or not allowed_file(file.filename):
            flash("Invalid file type! Only CSV, XLSX, or XLS files are allowed.", "danger")
            return render_template('ingredients/upload.html', current_page="ingredients")

        temp_path = None
        try:
            filename = secure_filename(file.filename)
            temp_path = os.path.join('uploads', filename)
            os.makedirs('uploads', exist_ok=True)
            file.save(temp_path)

            if filename.endswith('.csv'):
                df = pd.read_csv(temp_path)
            else:
                with pd.ExcelFile(temp_path) as excel_file:
                    if sheet_name not in excel_file.sheet_names:
                        flash(f"Sheet '{sheet_name}' not found in the Excel file. Available sheets: {', '.join(excel_file.sheet_names)}", "danger")
                        return render_template('ingredients/upload.html', current_page="ingredients")
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)

            df.columns = df.columns.str.strip()

            # Required columns for stocktake upload
            required_columns = ['Week Commencing', 'Item Code', 'SOH']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                flash(f"Missing required columns in file! Missing: {', '.join(missing_columns)}. Expected: {', '.join(required_columns)}", "danger")
                return render_template('ingredients/upload.html', current_page="ingredients")

            # Process each row
            created_count = 0
            updated_count = 0
            error_count = 0

            for _, row in df.iterrows():
                try:
                    # Extract required data
                    week_commencing_str = str(row['Week Commencing']).strip()
                    item_code = str(row['Item Code']).strip()
                    soh = float(row['SOH']) if pd.notnull(row['SOH']) else 0.0

                    # Parse week commencing date
                    try:
                        week_commencing = datetime.strptime(week_commencing_str, '%Y-%m-%d').date()
                    except ValueError:
                        try:
                            # Try alternative date formats
                            week_commencing = pd.to_datetime(week_commencing_str).date()
                        except:
                            raise ValueError(f"Invalid date format for Week Commencing: {week_commencing_str}")

                    # Check if item exists in Item Master as Raw Material
                    rm_type_id = get_rm_type_id()
                    if not rm_type_id:
                        error_count += 1
                        print(f"RM item type not found in system")
                        continue
                    
                    item = ItemMaster.query.filter_by(item_code=item_code, item_type_id=rm_type_id).first()
                    if not item:
                        error_count += 1
                        print(f"Item code '{item_code}' not found in Item Master as Raw Material")
                        continue

                    # Calculate order quantity using formula: if(SOH < MIN, MAX - SOH, 0)
                    min_level = item.min_level or 0
                    max_level = item.max_level or 0
                    order_quantity = max(0, max_level - soh) if soh < min_level else 0

                    # Calculate stock value based on UOM
                    price_per_kg = item.price_per_kg or 0
                    uom_name = item.uom.UOMName if item.uom else ''
                    
                    if uom_name.lower() in ['kg', 'liter']:
                        stock_value = soh * price_per_kg
                        price_uom = price_per_kg
                    else:
                        # For other UOMs, use price_per_kg as price_per_uom (could be enhanced)
                        price_uom = price_per_kg
                        stock_value = soh * price_uom

                    # Check if stocktake record already exists for this item and week
                    existing_stocktake = RawMaterialStocktake.query.filter_by(
                        item_code=item_code,
                        week_commencing=week_commencing
                    ).first()

                    if existing_stocktake:
                        # Update existing stocktake
                        existing_stocktake.current_stock = soh
                        existing_stocktake.order_quantity = order_quantity
                        existing_stocktake.price_uom = price_uom
                        existing_stocktake.stock_value = stock_value
                        existing_stocktake.category_id = item.category_id
                        updated_count += 1
                    else:
                        # Create new stocktake record
                        new_stocktake = RawMaterialStocktake(
                            week_commencing=week_commencing,
                            stocktake_type='weekly',  # Default to weekly
                            user='System Upload',     # Default user
                            item_code=item_code,
                            category_id=item.category_id,
                            current_stock=soh,
                            order_quantity=order_quantity,
                            price_uom=price_uom,
                            stock_value=stock_value,
                            notes=f"Uploaded from file: {filename}"
                        )
                        db.session.add(new_stocktake)
                        created_count += 1

                except Exception as e:
                    error_count += 1
                    print(f"Error processing row for item code {item_code}: {str(e)}")

            db.session.commit()
            flash(f"Stocktake upload completed! Created: {created_count}, Updated: {updated_count}, Errors: {error_count}", "success")
            return redirect(url_for('ingredients.ingredients_list'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error processing file: {str(e)}", "danger")
            return render_template('ingredients/upload.html', current_page="ingredients")

        finally:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except PermissionError as e:
                    print(f"Warning: Could not delete temporary file {temp_path}: {e}")

    return render_template('ingredients/upload.html', current_page="ingredients")

@ingredients_bp.route('/ingredients_download_excel', methods=['GET'])
def ingredients_download_excel():
    from app import db
    from models.item_master import ItemMaster
    from flask import make_response
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    try:
        # Get current filtered ingredients
        search_item_code = request.args.get('item_code', '').strip()
        search_description = request.args.get('description', '').strip()
        search_category = request.args.get('category', '').strip()

        ingredients_query = ItemMaster.query.filter(ItemMaster.item_type_id == get_rm_type_id())
        
        if search_item_code:
            ingredients_query = ingredients_query.filter(ItemMaster.item_code.ilike(f"%{search_item_code}%"))
        if search_description:
            ingredients_query = ingredients_query.filter(ItemMaster.description.ilike(f"%{search_description}%"))

        ingredients = ingredients_query.all()

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Ingredients"

        # Headers
        headers = ['Item Code', 'Description', 'Category', 'Department', 'UOM', 'Min Level', 'Max Level', 'Price per KG', 'Active']
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data
        for ingredient in ingredients:
            row_data = [
                ingredient.item_code,
                ingredient.description,
                ingredient.category.category_name if ingredient.category else '',
                ingredient.department.department_name if ingredient.department else '',
                ingredient.uom.UOM if ingredient.uom else '',
                ingredient.min_level or 0,
                ingredient.max_level or 0,
                ingredient.price_per_kg or 0,
                'Yes' if ingredient.is_active else 'No'
            ]
            ws.append(row_data)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = f'attachment; filename=ingredients_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    except Exception as e:
        flash(f"Error generating Excel file: {str(e)}", "danger")
        return redirect(url_for('ingredients.ingredients_list'))

@ingredients_bp.route('/ingredients_download_template', methods=['GET'])
def ingredients_download_template():
    from flask import make_response
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime, timedelta

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Stocktake Template"

        # Headers for stocktake upload
        headers = ['Week Commencing', 'Item Code', 'Description', 'Category', 'Department', 'SOH']
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add sample data with current and next week dates
        today = datetime.now().date()
        monday_this_week = today - timedelta(days=today.weekday())
        monday_next_week = monday_this_week + timedelta(days=7)

        sample_data = [
            [monday_this_week.strftime('%Y-%m-%d'), 'RM0001', 'Salt', 'Spices', 'Production', 25.50],
            [monday_this_week.strftime('%Y-%m-%d'), 'RM0002', 'Pepper Black Ground', 'Spices', 'Production', 12.75],
            [monday_next_week.strftime('%Y-%m-%d'), 'RM0001', 'Salt', 'Spices', 'Production', 18.25]
        ]

        for row_data in sample_data:
            ws.append(row_data)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add instructions sheet
        instructions_ws = wb.create_sheet("Instructions")
        instructions = [
            ["Field", "Description", "Required", "Example"],
            ["Week Commencing", "Date in YYYY-MM-DD format", "Yes", monday_this_week.strftime('%Y-%m-%d')],
            ["Item Code", "Must exist in Item Master", "Yes", "RM0001"],
            ["Description", "Auto-populated from Item Master", "No", "Salt"],
            ["Category", "Auto-populated from Item Master", "No", "Spices"],
            ["Department", "Auto-populated from Item Master", "No", "Production"],
            ["SOH", "Stock on Hand (numeric)", "Yes", "25.50"],
            ["", "", "", ""],
            ["Auto-calculated Fields:", "", "", ""],
            ["Order Quantity", "if(SOH < Min Level, Max Level - SOH, 0)", "Auto", ""],
            ["Stock Value", "Based on UOM and pricing from Item Master", "Auto", ""],
            ["Price/UOM", "From Item Master pricing", "Auto", ""],
            ["Min/Max Levels", "From Item Master", "Auto", ""],
            ["UOM", "From Item Master", "Auto", ""]
        ]

        for row in instructions:
            instructions_ws.append(row)

        # Style instructions header
        for cell in instructions_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Auto-adjust column widths for instructions
        for column in instructions_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            instructions_ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=stocktake_upload_template.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response

    except Exception as e:
        from flask import jsonify
        return jsonify({"error": f"Error generating template: {str(e)}"}), 500

@ingredients_bp.route('/autocomplete_ingredients', methods=['GET'])
def autocomplete_ingredients():
    from app import db
    from models.item_master import ItemMaster

    search = request.args.get('query', '').strip()

    if not search:
        return jsonify([])

    try:
        results = db.session.query(ItemMaster.item_code, ItemMaster.description).filter(
            ItemMaster.item_type_id == get_rm_type_id(),
            ItemMaster.item_code.ilike(f"{search}%")
        ).limit(10).all()
        
        suggestions = [{"item_code": row.item_code, "description": row.description} for row in results]
        return jsonify(suggestions)
    except Exception as e:
        print("Error fetching ingredient autocomplete suggestions:", e)
        return jsonify([])

@ingredients_bp.route('/get_item_details/<item_code>', methods=['GET'])
def get_item_details(item_code):
    from app import db
    from models.item_master import ItemMaster
    from models.category import Category
    from models.department import Department
    from models.uom import UOM

    try:
        rm_type_id = get_rm_type_id()
        if not rm_type_id:
            return jsonify({'error': 'RM item type not found in system'}), 400

        item = ItemMaster.query.filter_by(item_code=item_code, item_type_id=rm_type_id).first()
        if item:
            return jsonify({
                'id': item.id,
                'item_code': item.item_code,
                'description': item.description or '',
                'category_id': item.category_id,
                'department_id': item.department_id,
                'uom_id': item.uom_id,
                'min_level': item.min_level or 0,
                'max_level': item.max_level or 0,
                'price_per_kg': item.price_per_kg or 0,
                'price_per_uom': item.price_per_uom or 0,
                'is_active': item.is_active
            })
        else:
            return jsonify({'error': 'Item not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@ingredients_bp.route('/get_search_stocktakes', methods=['GET'])
def get_search_stocktakes():
    from app import db
    from models.raw_material_stocktake import RawMaterialStocktake
    from models.item_master import ItemMaster
    from models.category import Category

    search_item_code = request.args.get('item_code', '').strip()
    search_description = request.args.get('description', '').strip()
    search_category = request.args.get('category', '').strip()
    search_week_commencing = request.args.get('week_commencing', '').strip()
    sort_by = request.args.get('sort_by', 'id').strip()
    sort_direction = request.args.get('sort_direction', 'asc').strip()

    try:
        stocktakes_query = RawMaterialStocktake.query.join(ItemMaster, RawMaterialStocktake.item_code == ItemMaster.item_code)

        if search_item_code:
            stocktakes_query = stocktakes_query.filter(RawMaterialStocktake.item_code.ilike(f"%{search_item_code}%"))
        if search_description:
            stocktakes_query = stocktakes_query.filter(ItemMaster.description.ilike(f"%{search_description}%"))
        if search_category:
            stocktakes_query = stocktakes_query.join(Category).filter(Category.name.ilike(f"%{search_category}%"))
        if search_week_commencing:
            stocktakes_query = stocktakes_query.filter(RawMaterialStocktake.week_commencing == search_week_commencing)

        # Apply sorting
        if sort_by in ['item_code', 'week_commencing', 'stocktake_type', 'user', 'current_stock', 'order_quantity', 'price_uom', 'stock_value']:
            if sort_direction == 'desc':
                stocktakes_query = stocktakes_query.order_by(desc(getattr(RawMaterialStocktake, sort_by)))
            else:
                stocktakes_query = stocktakes_query.order_by(asc(getattr(RawMaterialStocktake, sort_by)))
        else:
            # Default sort by most recent
            stocktakes_query = stocktakes_query.order_by(desc(RawMaterialStocktake.created_at))

        stocktakes = stocktakes_query.all()

        stocktakes_data = []
        for stocktake in stocktakes:
            stocktakes_data.append({
                "id": stocktake.id,
                "week_commencing": stocktake.week_commencing_str,
                "stocktake_type": stocktake.stocktake_type or "",
                "user": stocktake.user or "",
                "item_code": stocktake.item_code or "",
                "description": stocktake.item.description if stocktake.item else "",
                "uom": stocktake.item.uom.UOMName if stocktake.item and stocktake.item.uom else "",
                "department": stocktake.item.department.departmentName if stocktake.item and stocktake.item.department else "",
                "current_stock": stocktake.current_stock if stocktake.current_stock is not None else "",
                "min_level": stocktake.item.min_level if stocktake.item and stocktake.item.min_level is not None else "",
                "max_level": stocktake.item.max_level if stocktake.item and stocktake.item.max_level is not None else "",
                "order_quantity": stocktake.order_quantity if stocktake.order_quantity is not None else "",
                "price_uom": stocktake.price_uom if stocktake.price_uom is not None else "",
                "price_kg": stocktake.item.price_per_kg if stocktake.item and stocktake.item.price_per_kg is not None else "",
                "stock_value": stocktake.stock_value if stocktake.stock_value is not None else "",
                "notes": stocktake.notes or ""
            })

        return jsonify(stocktakes_data)
    except Exception as e:
        print(f"Error fetching search stocktakes: {str(e)}")
        return jsonify([]), 500 