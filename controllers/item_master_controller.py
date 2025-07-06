from flask import Blueprint, render_template, request, jsonify, redirect, url_for, send_file, flash, session
from database import db
from models.item_master import ItemMaster
from models.category import Category
from models.department import Department
from models.machinery import Machinery
from models.uom import UOM
from models.allergen import Allergen
from models.item_type import ItemType
from models.user import User
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os
from werkzeug.utils import secure_filename

item_master_bp = Blueprint('item_master', __name__)

@item_master_bp.route('/item-master', methods=['GET'])
def item_master_list():
    # Get all lookup data for the filters
    categories = Category.query.all()
    departments = Department.query.all()
    item_types = ItemType.query.all()
    return render_template('item_master/list.html',
                         categories=categories,
                         departments=departments,
                         item_types=item_types,
                         current_page='item_master')

@item_master_bp.route('/item-master/create', methods=['GET'])
def item_master_create():
    # Get all lookup data
    categories = Category.query.all()
    departments = Department.query.all()
    machineries = Machinery.query.all()
    uoms = UOM.query.all()
    allergens = Allergen.query.all()
    item_types = ItemType.query.all()
    
    # Get WIP and WIPF items for component dropdowns
    wip_items = ItemMaster.query.join(ItemType).filter(ItemType.type_name == 'WIP').all()
    wipf_items = ItemMaster.query.join(ItemType).filter(ItemType.type_name == 'WIPF').all()
    
    return render_template('item_master/create.html',
                         categories=categories,
                         departments=departments,
                         machineries=machineries,
                         uoms=uoms,
                         allergens=allergens,
                         item_types=item_types,
                         wip_items=wip_items,
                         wipf_items=wipf_items,
                         current_page='item_master')

@item_master_bp.route('/item-master/edit/<int:id>', methods=['GET'])
def item_master_edit(id):
    # Get the item
    item = ItemMaster.query.get_or_404(id)
    
    # Get all lookup data
    categories = Category.query.all()
    departments = Department.query.all()
    machineries = Machinery.query.all()
    uoms = UOM.query.all()
    allergens = Allergen.query.all()
    item_types = ItemType.query.all()
    
    # Get WIP and WIPF items for component dropdowns
    wip_items = ItemMaster.query.join(ItemType).filter(ItemType.type_name == 'WIP').all()
    wipf_items = ItemMaster.query.join(ItemType).filter(ItemType.type_name == 'WIPF').all()

    return render_template('item_master/edit.html',
                         categories=categories,
                         departments=departments,
                         machineries=machineries,
                         uoms=uoms,
                         allergens=allergens,
                         item=item,
                         item_types=item_types,
                         wip_items=wip_items,
                         wipf_items=wipf_items,
                         current_page='item_master')

@item_master_bp.route('/get_items', methods=['GET'])
def get_items():
    search_code = request.args.get('item_code', '').strip()
    search_description = request.args.get('description', '').strip()
    search_type = request.args.get('item_type', '').strip()
    
    # Get sorting parameters
    sort_by = request.args.get('sort_by', 'item_code')  # Default sort by item_code
    sort_order = request.args.get('sort_order', 'asc')  # Default ascending
    
    # Start with base query and join ItemType for all queries
    query = ItemMaster.query.join(ItemType, ItemMaster.item_type_id == ItemType.id)
    
    # Apply filters
    if search_code:
        query = query.filter(ItemMaster.item_code.ilike(f"%{search_code}%"))
    if search_description:
        query = query.filter(ItemMaster.description.ilike(f"%{search_description}%"))
    if search_type:
        # Use exact type name matching
        query = query.filter(ItemType.type_name == search_type)

    # Apply sorting with proper joins
    if sort_by == 'item_code':
        query = query.order_by(ItemMaster.item_code.asc() if sort_order.lower() == 'asc' else ItemMaster.item_code.desc())
    elif sort_by == 'description':
        query = query.order_by(ItemMaster.description.asc() if sort_order.lower() == 'asc' else ItemMaster.description.desc())
    elif sort_by == 'item_type':
        # ItemType is already joined, so we can sort directly
        query = query.order_by(ItemType.type_name.asc() if sort_order.lower() == 'asc' else ItemType.type_name.desc())
    elif sort_by == 'category':
        query = query.outerjoin(Category).order_by(Category.name.asc() if sort_order.lower() == 'asc' else Category.name.desc())
    elif sort_by == 'department':
        query = query.outerjoin(Department).order_by(Department.departmentName.asc() if sort_order.lower() == 'asc' else Department.departmentName.desc())
    elif sort_by == 'machinery':
        query = query.outerjoin(Machinery).order_by(Machinery.machineryName.asc() if sort_order.lower() == 'asc' else Machinery.machineryName.desc())
    elif sort_by == 'uom':
        query = query.outerjoin(UOM).order_by(UOM.UOMName.asc() if sort_order.lower() == 'asc' else UOM.UOMName.desc())
    elif sort_by == 'min_level':
        query = query.order_by(ItemMaster.min_level.asc() if sort_order.lower() == 'asc' else ItemMaster.min_level.desc())
    elif sort_by == 'max_level':
        query = query.order_by(ItemMaster.max_level.asc() if sort_order.lower() == 'asc' else ItemMaster.max_level.desc())
    elif sort_by == 'price_per_kg':
        query = query.order_by(ItemMaster.price_per_kg.asc() if sort_order.lower() == 'asc' else ItemMaster.price_per_kg.desc())
    elif sort_by == 'price_per_uom':
        query = query.order_by(ItemMaster.price_per_uom.asc() if sort_order.lower() == 'asc' else ItemMaster.price_per_uom.desc())
    elif sort_by == 'supplier_name':
        query = query.order_by(ItemMaster.supplier_name.asc() if sort_order.lower() == 'asc' else ItemMaster.supplier_name.desc())
    else:
        # Default sorting
        query = query.order_by(ItemMaster.item_code.asc())

    items = query.all()
    
    items_data = []
    for item in items:
        item_data = {
            "id": item.id,
            "item_code": item.item_code,
            "description": item.description,
            "item_type": item.item_type.type_name if item.item_type else None,
            "category": item.category.name if item.category else None,
            "department": item.department.departmentName if item.department else None,
            "machinery": item.machinery.machineryName if item.machinery else None,
            "uom": item.uom.UOMName if item.uom else None,
            "min_level": item.min_level,
            "max_level": item.max_level,
            "price_per_kg": item.price_per_kg,
            "price_per_uom": item.price_per_uom,
            "calculation_factor": item.calculation_factor,
            "is_make_to_order": item.is_make_to_order,
            "kg_per_unit": item.kg_per_unit,
            "units_per_bag": item.units_per_bag,
            "avg_weight_per_unit": item.avg_weight_per_unit,
            "loss_percentage": item.loss_percentage,
            "supplier_name": item.supplier_name,
            "is_active": item.is_active,
            "allergens": [],  # Allergen relationship temporarily removed
            "created_by": None,  # User relationship temporarily removed  
            "updated_by": None,  # User relationship temporarily removed
            "created_at": item.created_at.strftime("%Y-%m-%d %H:%M:%S") if item.created_at else None,
            "updated_at": item.updated_at.strftime("%Y-%m-%d %H:%M:%S") if item.updated_at else None
        }
        items_data.append(item_data)
    
    return jsonify(items_data)

@item_master_bp.route('/item-master/create', methods=['POST'])
@item_master_bp.route('/item-master', methods=['POST'])
@item_master_bp.route('/item-master/edit/<int:id>', methods=['PUT'])
def save_item(id=None):
    try:
        data = request.get_json()
        
        # Get current user
        current_user_id = session.get('user_id')
        if not current_user_id:
            return jsonify({'success': False, 'message': 'User not authenticated'}), 401
        
        # Validate required fields
        required_fields = ['item_code', 'description', 'item_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field} is required'}), 400
        
        item_type_name = data.get('item_type')
        item_code = data.get('item_code')
        description = data.get('description')
        
        # Validate item type exists and get the ItemType object
        valid_item_type = ItemType.query.filter_by(type_name=item_type_name).first()
        if not valid_item_type:
            return jsonify({'success': False, 'message': f'Invalid item type: {item_type_name}'}), 400
        
        # For new item, check if item_code already exists
        if data.get('id') or id:
            item_id = data.get('id') or id
            item = ItemMaster.query.get_or_404(item_id)
            # For existing items, check if item_code is being changed and if new code already exists
            if item.item_code != item_code:
                if ItemMaster.query.filter_by(item_code=item_code).first():
                    return jsonify({'success': False, 'message': 'Item code already exists'}), 400
            # Update the updated_by field
            item.updated_by_id = current_user_id
        else:
            # Add "RM_" prefix for raw materials (only for new items)
            if item_type_name == 'Raw Material':
                item_code = f"{item_code}"
            if ItemMaster.query.filter_by(item_code=item_code).first():
                return jsonify({'success': False, 'message': 'Item code already exists'}), 400
            item = ItemMaster()
            # Set the created_by field for new items
            item.created_by_id = current_user_id
            item.updated_by_id = current_user_id
        
        # Update basic fields
        item.item_code = item_code
        item.description = description
        item.item_type_id = valid_item_type.id
        item.category_id = data.get('category_id') if data.get('category_id') else None
        item.department_id = data.get('department_id') if data.get('department_id') else None
        item.machinery_id = data.get('machinery_id') if data.get('machinery_id') else None
        item.uom_id = data.get('uom_id') if data.get('uom_id') else None
        item.min_level = data.get('min_level') if data.get('min_level') else None
        item.max_level = data.get('max_level') if data.get('max_level') else None
        item.price_per_kg = data.get('price_per_kg') if data.get('price_per_kg') else None
        item.price_per_uom = data.get('price_per_uom') if data.get('price_per_uom') else None
        item.calculation_factor = data.get('calculation_factor') if data.get('calculation_factor') else None
        item.supplier_name = data.get('supplier_name') if data.get('supplier_name') else None
        item.is_active = data.get('is_active', True)
        
        # Update type-specific fields
        if item_type_name == 'Raw Material':
            # Clear finished good fields
            item.is_make_to_order = False
            item.kg_per_unit = None
            item.units_per_bag = None
            item.avg_weight_per_unit = None
            item.loss_percentage = None
        else:
            item.is_make_to_order = data.get('is_make_to_order', False)
            item.kg_per_unit = data.get('kg_per_unit') if data.get('kg_per_unit') else None
            item.units_per_bag = data.get('units_per_bag') if data.get('units_per_bag') else None
            item.avg_weight_per_unit = data.get('avg_weight_per_unit') if data.get('avg_weight_per_unit') else None
            item.loss_percentage = data.get('loss_percentage') if data.get('loss_percentage') else None
        
        # Handle FG hierarchy relationships
        if item_type_name == 'FG':
            # Set WIP component relationship
            wip_item_id = data.get('wip_item_id')
            if wip_item_id:
                item.wip_item_id = int(wip_item_id)
            else:
                item.wip_item_id = None
                
            # Set WIPF component relationship  
            wipf_item_id = data.get('wipf_item_id')
            if wipf_item_id:
                item.wipf_item_id = int(wipf_item_id)
            else:
                item.wipf_item_id = None
        else:
            # Clear hierarchy relationships for non-FG items
            item.wip_item_id = None
            item.wipf_item_id = None
        
        # Handle allergens (temporarily disabled due to model changes)
        # if 'allergen_ids' in data:
        #     allergens = Allergen.query.filter(Allergen.allergens_id.in_(data['allergen_ids'])).all()
        #     item.allergens = allergens
        
        db.session.add(item)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Item saved successfully!', 'id': item.id}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@item_master_bp.route('/delete-item/<int:id>', methods=['DELETE'])
def delete_item(id):
    try:
        item = ItemMaster.query.get_or_404(id)
        db.session.delete(item)
        db.session.commit()
        return jsonify({'message': 'Item deleted successfully!'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@item_master_bp.route('/item-master/upload-excel', methods=['POST'])
def upload_excel():
    try:
        if 'excel_file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload a valid Excel file (.xlsx or .xls)'}), 400
        
        # Read the Excel file
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        # Expected columns: Item Code, Description, Type, Category, Department, UOM, Min Level, Max Level, Price Per Kg, Is Make To Order, Kg Per Unit, Units Per Bag, Loss Percentage, Is Active
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Validate required headers
        required_headers = ['Item Code', 'Description', 'Type']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_headers)}'}), 400
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Process each row (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):  # Skip empty rows
                    continue
                
                # Create dictionary from row data
                row_data = dict(zip(headers, row))
                
                # Skip if required fields are empty
                if not row_data.get('Item Code') or not row_data.get('Description') or not row_data.get('Type'):
                    error_count += 1
                    errors.append(f'Row {row_num}: Missing required fields')
                    continue
                
                item_code = str(row_data['Item Code']).strip()
                description = str(row_data['Description']).strip()
                item_type = str(row_data['Type']).strip()
                
                # Check if item already exists
                existing_item = ItemMaster.query.filter_by(item_code=item_code).first()
                if existing_item:
                    error_count += 1
                    errors.append(f'Row {row_num}: Item code {item_code} already exists')
                    continue
                
                # Validate item type
                valid_item_type = ItemType.query.filter_by(type_name=item_type).first()
                if not valid_item_type:
                    error_count += 1
                    errors.append(f'Row {row_num}: Invalid item type "{item_type}"')
                    continue
                
                # Create new item
                item = ItemMaster()
                item.item_code = item_code
                item.description = description
                item.item_type = item_type
                
                # Optional fields
                if row_data.get('Category'):
                    category = Category.query.filter_by(name=str(row_data['Category']).strip()).first()
                    if category:
                        item.category_id = category.id
                
                if row_data.get('Department'):
                    department = Department.query.filter_by(departmentName=str(row_data['Department']).strip()).first()
                    if department:
                        item.department_id = department.id
                
                if row_data.get('UOM'):
                    uom = UOM.query.filter_by(UOMName=str(row_data['UOM']).strip()).first()
                    if uom:
                        item.uom_id = uom.id
                
                # Text fields
                if row_data.get('Supplier Name'):
                    item.supplier_name = str(row_data['Supplier Name']).strip()
                
                # Numeric fields
                try:
                    if row_data.get('Min Level'):
                        item.min_level = float(row_data['Min Level'])
                    if row_data.get('Max Level'):
                        item.max_level = float(row_data['Max Level'])
                    if row_data.get('Price Per Kg'):
                        item.price_per_kg = float(row_data['Price Per Kg'])
                    if row_data.get('Price Per UOM'):
                        item.price_per_uom = float(row_data['Price Per UOM'])
                    if row_data.get('Kg Per Unit'):
                        item.kg_per_unit = float(row_data['Kg Per Unit'])
                    if row_data.get('Units Per Bag'):
                        item.units_per_bag = int(row_data['Units Per Bag'])
                    if row_data.get('Loss Percentage'):
                        item.loss_percentage = float(row_data['Loss Percentage'])
                except (ValueError, TypeError):
                    # If conversion fails, skip the numeric field
                    pass
                
                # Boolean fields
                if row_data.get('Is Make To Order'):
                    value = str(row_data['Is Make To Order']).strip().lower()
                    item.is_make_to_order = value in ['true', '1', 'yes', 'y']
                
                if row_data.get('Is Active'):
                    value = str(row_data['Is Active']).strip().lower()
                    item.is_active = value in ['true', '1', 'yes', 'y']
                else:
                    item.is_active = True  # Default to active
                
                db.session.add(item)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f'Row {row_num}: {str(e)}')
        
        db.session.commit()
        
        message = f'Upload completed: {success_count} items added successfully'
        if error_count > 0:
            message += f', {error_count} errors occurred'
            if len(errors) <= 5:  # Show first 5 errors
                message += f'. Errors: {"; ".join(errors)}'
        
        return jsonify({'message': message}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 500

@item_master_bp.route('/item-master/download-excel', methods=['GET'])
def download_excel():
    try:
        # Get search parameters
        search_code = request.args.get('item_code', '').strip()
        search_description = request.args.get('description', '').strip()
        search_type = request.args.get('item_type', '').strip()
        
        # Build query with same logic as get_items
        query = ItemMaster.query
        
        if search_code:
            query = query.filter(ItemMaster.item_code.ilike(f"%{search_code}%"))
        if search_description:
            query = query.filter(ItemMaster.description.ilike(f"%{search_description}%"))
        if search_type:
            item_type_obj = ItemType.query.filter_by(type_name=search_type).first()
            if item_type_obj:
                query = query.filter(ItemMaster.item_type_id == item_type_obj.id)
        
        items = query.all()
        
        # Create workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Item Master"
        
        # Define headers
        headers = [
            'Item Code', 'Description', 'Type', 'Category', 'Department', 
            'UOM', 'Min Level', 'Max Level', 'Price Per Kg', 'Price Per UOM', 
            'Supplier Name', 'Is Make To Order', 'Kg Per Unit', 'Units Per Bag', 
            'Loss Percentage', 'Is Active'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, item in enumerate(items, 2):
            sheet.cell(row=row, column=1, value=item.item_code)
            sheet.cell(row=row, column=2, value=item.description)
            sheet.cell(row=row, column=3, value=item.item_type.type_name if item.item_type else '')
            sheet.cell(row=row, column=4, value=item.category.name if item.category else '')
            sheet.cell(row=row, column=5, value=item.department.departmentName if item.department else '')
            sheet.cell(row=row, column=6, value=item.uom.UOMName if item.uom else '')
            sheet.cell(row=row, column=7, value=item.min_level)
            sheet.cell(row=row, column=8, value=item.max_level)
            sheet.cell(row=row, column=9, value=item.price_per_kg)
            sheet.cell(row=row, column=10, value=item.price_per_uom)
            sheet.cell(row=row, column=11, value=item.supplier_name or '')
            sheet.cell(row=row, column=12, value='Yes' if item.is_make_to_order else 'No')
            sheet.cell(row=row, column=13, value=item.kg_per_unit)
            sheet.cell(row=row, column=14, value=item.units_per_bag)
            sheet.cell(row=row, column=15, value=item.loss_percentage)
            sheet.cell(row=row, column=16, value='Yes' if item.is_active else 'No')
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'item_master_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate Excel file: {str(e)}'}), 500

@item_master_bp.route('/item-master/download-template', methods=['GET'])
def download_template():
    try:
        # Create workbook with template structure
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Item Master Template"
        
        # Define headers with descriptions
        headers = [
            'Item Code', 'Description', 'Type', 'Category', 'Department', 
            'UOM', 'Min Level', 'Max Level', 'Price Per Kg', 'Price Per UOM', 
            'Supplier Name', 'Is Make To Order', 'Kg Per Unit', 'Units Per Bag', 
            'Loss Percentage', 'Is Active'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data row
        sample_data = [
            'ITEM001', 'Sample Item Description', 'Raw Material', 'Category Name', 'Department Name',
            'KG', '10', '100', '5.50', '6.00', 'ABC Suppliers Ltd', 'No', '', '', '', 'Yes'
        ]
        
        for col, value in enumerate(sample_data, 1):
            sheet.cell(row=2, column=col, value=value)
        
        # Add instructions in a separate sheet
        instructions_sheet = workbook.create_sheet("Instructions")
        instructions = [
            "ITEM MASTER UPLOAD INSTRUCTIONS",
            "",
            "Required Columns:",
            "- Item Code: Unique identifier for the item",
            "- Description: Item description",
            "- Type: Must match existing item types in the system",
            "",
            "Optional Columns:",
            "- Category: Must match existing categories",
            "- Department: Must match existing departments",
            "- UOM: Unit of measure (must match existing UOMs)",
            "- Min Level: Minimum stock level (numeric)",
            "- Max Level: Maximum stock level (numeric)",
            "- Price Per Kg: Price per kilogram (numeric, for Raw Materials)",
            "- Price Per UOM: Price per unit of measure (numeric)",
            "- Supplier Name: Name of the supplier for this item",
            "- Is Make To Order: Yes/No (for Finished Goods)",
            "- Kg Per Unit: Kilograms per unit (numeric, for Finished Goods)",
            "- Units Per Bag: Units per bag (numeric, for Finished Goods)",
            "- Loss Percentage: Loss percentage (numeric, for Finished Goods)",
            "- Is Active: Yes/No (defaults to Yes if not specified)",
            "",
            "Notes:",
            "- Do not modify the header row",
            "- Empty rows will be skipped",
            "- Items with duplicate codes will be rejected",
            "- Invalid references (category, department, etc.) will be ignored",
            "- Boolean fields accept: Yes/No, True/False, 1/0, Y/N"
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = instructions_sheet.cell(row=row, column=1, value=instruction)
            if row == 1:  # Title
                cell.font = Font(bold=True, size=14)
            elif instruction.endswith(":"):  # Section headers
                cell.font = Font(bold=True)
        
        # Auto-adjust column widths for both sheets
        for sheet_obj in [sheet, instructions_sheet]:
            for column in sheet_obj.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                sheet_obj.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='item_master_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate template: {str(e)}'}), 500

@item_master_bp.route('/item-master/test-form', methods=['GET'])
def test_form():
    """Serve a simple test form for debugging form submission issues."""
    import os
    test_file_path = os.path.join(os.getcwd(), 'test_form_simple.html')
    if os.path.exists(test_file_path):
        with open(test_file_path, 'r', encoding='utf-8') as f:
            return f.read()
    else:
        return "Test form file not found", 404

@item_master_bp.route('/item-master/upload', methods=['GET', 'POST'])
def item_master_upload():
    """Render the upload page for Item Master Excel/CSV upload and handle file uploads."""
    if request.method == 'POST':
        # Handle file upload - reuse the existing upload_excel logic
        try:
            if 'file' not in request.files:
                flash('No file uploaded', 'error')
                return render_template('item_master/upload.html', current_page='item_master')
            
            file = request.files['file']
            if file.filename == '':
                flash('No file selected', 'error')
                return render_template('item_master/upload.html', current_page='item_master')
            
            if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
                flash('Please upload a valid Excel or CSV file (.xlsx, .xls, .csv)', 'error')
                return render_template('item_master/upload.html', current_page='item_master')
            
            # Process the file using existing upload logic
            # Read the Excel file
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            # Expected columns: Item Code, Description, Type, Category, Department, UOM, Min Level, Max Level, Price Per Kg, Is Make To Order, Kg Per Unit, Units Per Bag, Loss Percentage, Is Active
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Validate required headers
            required_headers = ['Item Code', 'Description', 'Type']
            missing_headers = [h for h in required_headers if h not in headers]
            if missing_headers:
                flash(f'Missing required columns: {", ".join(missing_headers)}', 'error')
                return render_template('item_master/upload.html', current_page='item_master')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Process each row (skip header)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # Skip empty rows
                        continue
                    
                    # Create dictionary from row data
                    row_data = dict(zip(headers, row))
                    
                    # Skip if required fields are empty
                    if not row_data.get('Item Code') or not row_data.get('Description') or not row_data.get('Type'):
                        error_count += 1
                        errors.append(f'Row {row_num}: Missing required fields')
                        continue
                    
                    item_code = str(row_data['Item Code']).strip()
                    description = str(row_data['Description']).strip()
                    item_type = str(row_data['Type']).strip()
                    
                    # Check if item already exists
                    existing_item = ItemMaster.query.filter_by(item_code=item_code).first()
                    if existing_item:
                        error_count += 1
                        errors.append(f'Row {row_num}: Item code {item_code} already exists')
                        continue
                    
                    # Validate item type
                    valid_item_type = ItemType.query.filter_by(type_name=item_type).first()
                    if not valid_item_type:
                        error_count += 1
                        errors.append(f'Row {row_num}: Invalid item type "{item_type}"')
                        continue
                    
                    # Create new item
                    item = ItemMaster()
                    item.item_code = item_code
                    item.description = description
                    # Find ItemType by name and set the foreign key
                    valid_item_type = ItemType.query.filter_by(type_name=item_type).first()
                    if valid_item_type:
                        item.item_type_id = valid_item_type.id
                    else:
                        # Skip this row if item type is invalid
                        error_count += 1
                        continue
                    
                    # Optional fields
                    if row_data.get('Category'):
                        category = Category.query.filter_by(name=str(row_data['Category']).strip()).first()
                        if category:
                            item.category_id = category.id
                    
                    if row_data.get('Department'):
                        department = Department.query.filter_by(departmentName=str(row_data['Department']).strip()).first()
                        if department:
                            item.department_id = department.id
                    
                    if row_data.get('UOM'):
                        uom = UOM.query.filter_by(UOMName=str(row_data['UOM']).strip()).first()
                        if uom:
                            item.uom_id = uom.id
                    
                    # Text fields
                    if row_data.get('Supplier Name'):
                        item.supplier_name = str(row_data['Supplier Name']).strip()
                    
                    # Numeric fields
                    try:
                        if row_data.get('Min Level'):
                            item.min_level = float(row_data['Min Level'])
                        if row_data.get('Max Level'):
                            item.max_level = float(row_data['Max Level'])
                        if row_data.get('Price Per Kg'):
                            item.price_per_kg = float(row_data['Price Per Kg'])
                        if row_data.get('Price Per UOM'):
                            item.price_per_uom = float(row_data['Price Per UOM'])
                        if row_data.get('Kg Per Unit'):
                            item.kg_per_unit = float(row_data['Kg Per Unit'])
                        if row_data.get('Units Per Bag'):
                            item.units_per_bag = int(row_data['Units Per Bag'])
                        if row_data.get('Loss Percentage'):
                            item.loss_percentage = float(row_data['Loss Percentage'])
                    except (ValueError, TypeError):
                        # If conversion fails, skip the numeric field
                        pass
                    
                    # Boolean fields
                    if row_data.get('Is Make To Order'):
                        value = str(row_data['Is Make To Order']).strip().lower()
                        item.is_make_to_order = value in ['true', '1', 'yes', 'y']
                    
                    if row_data.get('Is Active'):
                        value = str(row_data['Is Active']).strip().lower()
                        item.is_active = value in ['true', '1', 'yes', 'y']
                    else:
                        item.is_active = True  # Default to active
                    
                    db.session.add(item)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Row {row_num}: {str(e)}')
            
            db.session.commit()
            
            message = f'Upload completed: {success_count} items added successfully'
            if error_count > 0:
                message += f', {error_count} errors occurred'
                if len(errors) <= 5:  # Show first 5 errors
                    message += f'. Errors: {"; ".join(errors)}'
                flash(message, 'warning')
            else:
                flash(message, 'success')
            
            return render_template('item_master/upload.html', current_page='item_master')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Failed to process file: {str(e)}', 'error')
            return render_template('item_master/upload.html', current_page='item_master')
    
    # GET request - just render the upload page
    return render_template('item_master/upload.html', current_page='item_master')

@item_master_bp.route('/item-master/get-items-by-type/<item_type>', methods=['GET'])
def get_items_by_type(item_type):
    """Get items by item type for hierarchy selection"""
    try:
        # Check if user is authenticated
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Authentication required'}), 401
        
        # Get items by type
        items = ItemMaster.query.join(ItemType).filter(
            ItemType.type_name == item_type,
            ItemMaster.is_active == True
        ).order_by(ItemMaster.item_code).all()
        
        items_data = []
        for item in items:
            items_data.append({
                'id': item.id,
                'item_code': item.item_code,
                'description': item.description or ''
            })
        
        return jsonify({'success': True, 'items': items_data})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@item_master_bp.route('/autocomplete-item-code', methods=['GET'])
def autocomplete_item_code():
    query = request.args.get('query', '').strip()
    if not query:
        return jsonify([])
    
    # Search for items matching the query
    items = ItemMaster.query.filter(
        ItemMaster.item_code.ilike(f"%{query}%")
    ).limit(50).all()
    
    suggestions = []
    for item in items:
        suggestions.append({
            'item_code': item.item_code,
            'description': item.description,
            'item_type': item.item_type.type_name if item.item_type else None
        })
    
    return jsonify(suggestions)

@item_master_bp.route('/search_item_codes', methods=['GET'])
def search_item_codes():
    # Check if user is authenticated
    if 'user_id' not in session:
        return jsonify([]), 401
    
    term = request.args.get('term', '')
    if not term or len(term) < 2:
        return jsonify([])
    
    try:
        # Search for item codes that match the term
        items = ItemMaster.query.filter(ItemMaster.item_code.ilike(f'%{term}%')).limit(10).all()
        
        # Return list of matching item codes with descriptions
        results = [{
            'item_code': item.item_code,
            'description': item.description or ''
        } for item in items]
        
        return jsonify(results)
    except Exception as e:
        return jsonify([]), 500