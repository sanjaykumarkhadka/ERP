from flask import Blueprint, render_template, request, redirect, send_file, url_for, flash, jsonify
from sqlalchemy.sql import text
from decimal import Decimal
import sqlalchemy.exc
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
from database import db
from models import Production, RecipeMaster, UsageReport, RawMaterialReport, ItemMaster, ItemType
from models.usage_report import UsageReport
from models.recipe_master import RecipeMaster
from models.production import Production
# from models.joining import Joining  # REMOVED - joining table deprecated
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# Create a Blueprint for recipe routes
recipe_bp = Blueprint('recipe', __name__, template_folder='templates')

def get_monday_date(date_str):
    """Convert any date to the previous Monday if it's not already a Monday.
    Supports both YYYY-MM-DD and DD/MM/YYYY formats."""
    try:
        # Try YYYY-MM-DD format first
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        try:
            # Try DD/MM/YYYY format
            date = datetime.strptime(date_str, '%d/%m/%Y').date()
        except ValueError:
            raise ValueError("Date must be in YYYY-MM-DD or DD/MM/YYYY format")
    
    days_since_monday = date.weekday()  # Monday = 0, Sunday = 6
    monday = date - timedelta(days=days_since_monday)
    return monday

@recipe_bp.route('/recipe', methods=['GET', 'POST'])
def recipe_page():
    if request.method == 'POST':
        try:
            data = request.get_json() 
            recipes_data = data.get('recipes', [])

            if not recipes_data:
                return jsonify({'error': 'No recipes data provided.'}), 400
            
            # Validate descriptions are unique
            descriptions = {recipe.get('description') for recipe in recipes_data}
            if len(descriptions) > 1:
                return jsonify({'error': 'All recipes must have the same description.'}), 400

            # Process all recipes
            for recipe_data in recipes_data:
                recipe_id = recipe_data.get('recipe_id')
                recipe_code = recipe_data.get('recipe_code')
                description = recipe_data.get('description')
                recipe_wip_id = recipe_data.get('recipe_wip_id')  # Changed from finished_good_id
                component_item_id = recipe_data.get('component_item_id')  # Changed from raw_material_id
                kg_per_batch = recipe_data.get('kg_per_batch')

                if not all([recipe_code, description, recipe_wip_id, component_item_id, kg_per_batch]):
                    return jsonify({'error': 'Required fields are missing.'}), 400
                
                # Validate that recipe_wip_id is a WIP item
                wip_item = db.session.query(ItemMaster).join(
                    ItemType, ItemMaster.item_type_id == ItemType.id
                ).filter(
                    ItemMaster.id == recipe_wip_id,
                    ItemType.type_name == 'WIP'
                ).first()
                
                if not wip_item:
                    return jsonify({'error': 'Recipe code must be a WIP item.'}), 400
                
                # Validate kg_per_batch is a number
                try:
                    kg_per_batch = Decimal(kg_per_batch)
                    if kg_per_batch <= 0:
                        return jsonify({'error': 'Kg per batch cannot be negative or zero.'}), 400
                except (ValueError, TypeError):
                    return jsonify({'error': 'Invalid kg per batch value.'}), 400

                if recipe_id:  # Edit case
                    recipe = RecipeMaster.query.get_or_404(recipe_id)
                    # Update using actual database field names
                    recipe.recipe_wip_id = recipe_wip_id
                    recipe.component_item_id = component_item_id
                    recipe.quantity_kg = kg_per_batch  # Use actual database field name
                else:  # Add case
                    recipe = RecipeMaster(
                        recipe_wip_id=recipe_wip_id,
                        component_item_id=component_item_id,
                        quantity_kg=kg_per_batch  # Use actual database field name
                    )
                    db.session.add(recipe)

                db.session.flush()

            # Recalculate percentages for all recipes with the same recipe_wip_id
            recipe_wip_id = recipes_data[0]['recipe_wip_id']
            recipes_to_update = RecipeMaster.query.filter(
                RecipeMaster.recipe_wip_id == recipe_wip_id
            ).all()

            total_quantity = sum(float(r.quantity_kg) for r in recipes_to_update)
            # Note: percentage is calculated automatically in get_search_recipes, not stored

            db.session.commit()
            return jsonify({'message': 'Recipes saved successfully!'}), 200

        except sqlalchemy.exc.IntegrityError as e:
            db.session.rollback()
            return jsonify({'error': 'Database error: Duplicate entry or invalid data.'}), 400
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'An error occurred: {str(e)}'}), 500

    # GET request: render the page
    search_recipe_code = request.args.get('recipe_code', '')
    search_description = request.args.get('description', '')
    edit_id = request.args.get('edit_id')
    
    # Get only WIP items for recipe codes and RM items for components
    wip_items = db.session.query(ItemMaster).join(
        ItemType, ItemMaster.item_type_id == ItemType.id
    ).filter(ItemType.type_name == 'WIP').order_by(ItemMaster.item_code).all()
    
    component_items = db.session.query(ItemMaster).join(
        ItemType, ItemMaster.item_type_id == ItemType.id  
    ).filter(ItemType.type_name == 'RM').order_by(ItemMaster.item_code).all()

    return render_template('recipe/recipe.html', 
                         search_recipe_code=search_recipe_code,
                         search_description=search_description,
                         recipes=RecipeMaster.query.all(),
                         wip_items=wip_items,  # Only WIP items for recipe codes
                         component_items=component_items,  # Only RM items for components
                         current_page='recipe')

@recipe_bp.route('/recipe/delete/<int:id>', methods=['POST'])
def delete_recipe(id):
    try:
        recipe = RecipeMaster.query.get_or_404(id)
        recipe_wip_id = recipe.recipe_wip_id
        db.session.delete(recipe)
        db.session.commit()

        # Recalculate percentages for remaining recipes with the same recipe_wip_id
        recipes_to_update = RecipeMaster.query.filter(RecipeMaster.recipe_wip_id == recipe_wip_id).all()
        if recipes_to_update:
            total_quantity = sum(float(r.quantity_kg) for r in recipes_to_update)
            # Note: percentage is calculated automatically in get_search_recipes, not stored

        return jsonify({'message': 'Recipe deleted successfully!'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500

@recipe_bp.route('/autocomplete_recipe', methods=['GET'])
def autocomplete_recipe():
    search = request.args.get('query', '').strip()
    if not search:
        return jsonify([])
    try:
        # Only show WIP items as recipe codes (as requested by user)
        from sqlalchemy.orm import aliased
        WipItem = aliased(ItemMaster)
        
        query = db.session.query(
            WipItem.item_code.label('recipe_code'),
            WipItem.description.label('description')
        ).join(
            RecipeMaster,
            RecipeMaster.recipe_wip_id == WipItem.id
        ).join(
            ItemType,
            WipItem.item_type_id == ItemType.id
        ).filter(
            ItemType.type_name == 'WIP',  # Only WIP items
            WipItem.item_code.ilike(f"{search}%")
        ).distinct().order_by(WipItem.item_code).limit(25)
        
        results = query.all()
        suggestions = [
            {
                "recipe_code": row.recipe_code, 
                "description": row.description
            } 
            for row in results
        ]
        return jsonify(suggestions)
    except Exception as e:
        print(f"Error fetching recipe autocomplete suggestions: {e}")
        return jsonify([])

@recipe_bp.route('/get_search_recipes', methods=['GET'])
def get_search_recipes():
    search_recipe_code = request.args.get('recipe_code', '').strip()
    search_description = request.args.get('description', '').strip()
    
    # Create aliases for the ItemMaster joins
    from sqlalchemy.orm import aliased
    from sqlalchemy import func
    RawMaterialItem = aliased(ItemMaster)
    WipItem = aliased(ItemMaster)  # WIP items are recipe codes
    
    # Join with both raw material and WIP items using correct field names
    recipes_query = db.session.query(
        RecipeMaster,
        RawMaterialItem.item_code.label('raw_material_code'),
        RawMaterialItem.description.label('raw_material'),
        WipItem.item_code.label('wip_code'),  # WIP item code is the recipe code
        WipItem.description.label('wip_description')  # WIP item description
    ).join(
        RawMaterialItem,
        RecipeMaster.component_item_id == RawMaterialItem.id
    ).join(
        WipItem,
        RecipeMaster.recipe_wip_id == WipItem.id
    )
    
    # Apply filters based on search criteria (only filter by WIP items)
    if search_recipe_code:
        recipes_query = recipes_query.filter(WipItem.item_code.ilike(f"%{search_recipe_code}%"))
    if search_description:
        recipes_query = recipes_query.filter(WipItem.description.ilike(f"%{search_description}%"))
    
    recipes = recipes_query.all()
    
    # Calculate total quantities for each recipe to compute percentages
    recipe_totals = {}
    for recipe in recipes:
        recipe_wip_id = recipe.RecipeMaster.recipe_wip_id
        if recipe_wip_id not in recipe_totals:
            # Get total quantity for this recipe
            total_query = db.session.query(
                func.sum(RecipeMaster.quantity_kg).label('total_kg')
            ).filter(RecipeMaster.recipe_wip_id == recipe_wip_id).first()
            recipe_totals[recipe_wip_id] = float(total_query.total_kg) if total_query.total_kg else 0.0
    
    recipes_data = []
    for recipe in recipes:
        # Recipe code and description come from the WIP item
        recipe_code = recipe.wip_code
        description = recipe.wip_description
        
        # Calculate percentage: (component quantity / total recipe quantity) * 100
        component_kg = float(recipe.RecipeMaster.quantity_kg) if recipe.RecipeMaster.quantity_kg else 0.0
        total_kg = recipe_totals.get(recipe.RecipeMaster.recipe_wip_id, 0.0)
        percentage = (component_kg / total_kg * 100) if total_kg > 0 else 0.0
        
        recipes_data.append({
            "id": recipe.RecipeMaster.id,
            "recipe_code": recipe_code,
            "description": description,
            "raw_material_code": recipe.raw_material_code,
            "raw_material": recipe.raw_material,
            "component_item_id": recipe.RecipeMaster.component_item_id,  # Use correct field name
            "recipe_wip_id": recipe.RecipeMaster.recipe_wip_id,  # Use correct field name
            "kg_per_batch": component_kg,
            "percentage": round(percentage, 2),  # Automatically calculated percentage
        })
    
    return jsonify(recipes_data)

@recipe_bp.route('/usage')
def usage():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    try:
        grouped_usage_data = {}
        
        # If no date filters provided, get data from usage_report table with stored percentages
        if not from_date or not to_date:
            # Get existing data from usage_report table - use stored percentages
            usage_reports = UsageReport.query.order_by(UsageReport.production_date.desc()).all()
            
            for report in usage_reports:
                date = report.production_date
                
                if date not in grouped_usage_data:
                    grouped_usage_data[date] = []
                    
                grouped_usage_data[date].append({
                    'week_commencing': report.week_commencing.strftime('%Y-%m-%d'),
                    'production_date': report.production_date.strftime('%Y-%m-%d'),
                    'production_code': 'N/A',  # Not stored in usage_report table
                    'recipe_code': report.recipe_code,
                    'component_material': report.raw_material,
                    'usage_kg': report.usage_kg,
                    'kg_per_batch': 0.0,  # Not stored in usage_report table
                    'percentage': report.percentage  # Use stored percentage from database
                })
        else:
            # Date filters provided - recalculate and save new data
            # Create aliases for ItemMaster to avoid conflicts
            ProductionItem = ItemMaster.__table__.alias('production_item')
            ComponentItem = ItemMaster.__table__.alias('component_item')
            
            # Query to get production and recipe usage data using new schema
            query = db.session.query(
                Production,
                RecipeMaster,
                ComponentItem.c.description.label('component_name')
            ).join(
                ProductionItem, Production.item_id == ProductionItem.c.id  # Join Production to ItemMaster (WIP item)
            ).join(
                RecipeMaster, ProductionItem.c.id == RecipeMaster.recipe_wip_id  # Join to RecipeMaster via recipe_wip_id
            ).join(
                ComponentItem, RecipeMaster.component_item_id == ComponentItem.c.id  # Join to component ItemMaster
            )
            
            # Apply date filters
            query = query.filter(
                Production.production_date >= from_date,
                Production.production_date <= to_date
            )
                
            # Clear existing usage_report data for the date range
            UsageReport.query.filter(
                UsageReport.production_date >= from_date,
                UsageReport.production_date <= to_date
            ).delete()
            
            # Get the results
            usage_data = query.all()
            
            # First pass: calculate recipe totals for percentage calculation
            recipe_totals = {}
            production_data = []
            
            for production, recipe, component_name in usage_data:
                date = production.production_date
                week_commencing = get_monday_date(date.strftime('%Y-%m-%d'))
                usage_kg = float(recipe.quantity_kg) * (production.batches or 0)
                recipe_code = production.item.item_code if production.item else 'Unknown'
                
                recipe_key = f"{date}_{recipe_code}"
                if recipe_key not in recipe_totals:
                    recipe_totals[recipe_key] = 0.0
                recipe_totals[recipe_key] += usage_kg
                
                production_data.append({
                    'production': production,
                    'recipe': recipe,
                    'component_name': component_name,
                    'date': date,
                    'week_commencing': week_commencing,
                    'usage_kg': usage_kg,
                    'recipe_code': recipe_code
                })
            
            # Second pass: save data with calculated percentages
            for data in production_data:
                recipe_key = f"{data['date']}_{data['recipe_code']}"
                total_kg = recipe_totals[recipe_key]
                percentage = (data['usage_kg'] / total_kg * 100) if total_kg > 0 else 0.0
                
                # Save to usage_report table
                usage_report = UsageReport(
                    week_commencing=data['week_commencing'],
                    production_date=data['date'],
                    recipe_code=data['recipe_code'],
                    raw_material=data['component_name'],
                    usage_kg=data['usage_kg'],
                    percentage=percentage,
                    created_at=datetime.now()
                )
                db.session.add(usage_report)
                
                # Group data for display
                if data['date'] not in grouped_usage_data:
                    grouped_usage_data[data['date']] = []
                    
                grouped_usage_data[data['date']].append({
                    'week_commencing': data['week_commencing'].strftime('%Y-%m-%d'),
                    'production_date': data['date'].strftime('%Y-%m-%d'),
                    'production_code': data['production'].production_code,
                    'recipe_code': data['recipe_code'],
                    'component_material': data['component_name'],
                    'usage_kg': data['usage_kg'],
                    'kg_per_batch': float(data['recipe'].quantity_kg),
                    'percentage': percentage
                })
            
            # Commit the data to usage_report table
            db.session.commit()
        
        return render_template('recipe/usage.html',
                             grouped_usage_data=grouped_usage_data,
                             from_date=from_date,
                             to_date=to_date,
                             current_page='usage')
                             
    except Exception as e:
        db.session.rollback()
        flash(f"An error occurred: {str(e)}", 'error')
        return render_template('recipe/usage.html',
                             grouped_usage_data={},
                             from_date=from_date,
                             to_date=to_date,
                             current_page='usage')

@recipe_bp.route('/usage/download')
def usage_download():
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    
    # Create aliases for ItemMaster to avoid conflicts
    ProductionItem = ItemMaster.__table__.alias('production_item')
    ComponentItem = ItemMaster.__table__.alias('component_item')
    
    # Query to get production and recipe usage data using new schema
    query = db.session.query(
        Production,
        RecipeMaster,
        ComponentItem.c.description.label('component_name')
    ).join(
        ProductionItem, Production.item_id == ProductionItem.c.id  # Join Production to ItemMaster (WIP item)
    ).join(
        RecipeMaster, ProductionItem.c.id == RecipeMaster.recipe_wip_id  # Join to RecipeMaster via recipe_wip_id
    ).join(
        ComponentItem, RecipeMaster.component_item_id == ComponentItem.c.id  # Join to component ItemMaster
    )
    
    # Apply date filters if provided
    if from_date and to_date:
        query = query.filter(
            Production.production_date >= from_date,
            Production.production_date <= to_date
        )
    
    # Get the results
    usage_data = query.all()
    
    # Create Excel file
    data = []
    for production, recipe, component_name in usage_data:
        # Calculate the Monday of the week for the production_date
        week_commencing = get_monday_date(production.production_date.strftime('%Y-%m-%d'))
        data.append({
            'Week Commencing': week_commencing.strftime('%Y-%m-%d'),
            'Production Date': production.production_date.strftime('%Y-%m-%d'),
            'Production Code': production.production_code,
            'Recipe Code': production.item.item_code if production.item else 'Unknown',  # Use actual item code
            'Component Material': component_name,
            'Usage Kg': float(recipe.quantity_kg) * (production.batches or 0),  # Calculate actual usage
            'Kg per Batch': float(recipe.quantity_kg)  # Use quantity_kg
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Usage Report', index=False)
    
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'usage_report_{from_date}_{to_date}.xlsx' if from_date and to_date else 'usage_report.xlsx'
    )

@recipe_bp.route('/recipe/upload-excel', methods=['POST'])
def upload_recipe_excel():
    try:
        if 'excel_file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['excel_file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Please upload a valid Excel file (.xlsx or .xls)'}), 400
        
        # Read the Excel file
        import openpyxl
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        # Expected columns: Recipe Code, Description, Finished Good Code, Raw Material Code, Kg Per Batch, UOM, Is Active
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Validate required headers
        required_headers = ['Recipe Code', 'Description', 'Finished Good Code', 'Raw Material Code', 'Kg Per Batch']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_headers)}'}), 400
        
        success_count = 0
        error_count = 0
        errors = []
        
        # Process each row (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):  # Skip empty rows
                    continue
                
                # Create dictionary from row data
                row_data = dict(zip(headers, row))
                
                # Skip if required fields are empty
                if not all([row_data.get('Recipe Code'), row_data.get('Description'), 
                           row_data.get('Finished Good Code'), row_data.get('Raw Material Code'), 
                           row_data.get('Kg Per Batch')]):
                    error_count += 1
                    errors.append(f'Row {row_num}: Missing required fields')
                    continue
                
                recipe_code = str(row_data['Recipe Code']).strip()
                description = str(row_data['Description']).strip()
                finished_good_code = str(row_data['Finished Good Code']).strip()
                raw_material_code = str(row_data['Raw Material Code']).strip()
                
                # Validate kg_per_batch is a number
                try:
                    kg_per_batch = float(row_data['Kg Per Batch'])
                    if kg_per_batch <= 0:
                        error_count += 1
                        errors.append(f'Row {row_num}: Kg per batch must be greater than 0')
                        continue
                except (ValueError, TypeError):
                    error_count += 1
                    errors.append(f'Row {row_num}: Invalid kg per batch value')
                    continue
                
                # Find finished good item
                finished_good = ItemMaster.query.filter_by(item_code=finished_good_code).first()
                if not finished_good:
                    error_count += 1
                    errors.append(f'Row {row_num}: Finished good code "{finished_good_code}" not found')
                    continue
                
                # Find raw material item
                raw_material = ItemMaster.query.filter_by(item_code=raw_material_code).first()
                if not raw_material:
                    error_count += 1
                    errors.append(f'Row {row_num}: Raw material code "{raw_material_code}" not found')
                    continue
                
                # Check if recipe already exists
                existing_recipe = RecipeMaster.query.filter(
                    RecipeMaster.finished_good_id == finished_good.id,
                    RecipeMaster.raw_material_id == raw_material.id
                ).first()
                if existing_recipe:
                    error_count += 1
                    errors.append(f'Row {row_num}: Recipe for {finished_good_code} -> {raw_material_code} already exists')
                    continue
                
                # Find UOM if specified
                uom_id = None
                if row_data.get('UOM'):
                    from models.uom import UOM
                    uom = UOM.query.filter_by(UOMName=str(row_data['UOM']).strip()).first()
                    if uom:
                        uom_id = uom.UOMID
                
                # Create new recipe
                recipe = RecipeMaster()
                recipe.recipe_code = recipe_code
                recipe.description = description
                recipe.finished_good_id = finished_good.id
                recipe.raw_material_id = raw_material.id
                recipe.kg_per_batch = kg_per_batch
                recipe.quantity_uom_id = uom_id
                
                db.session.add(recipe)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f'Row {row_num}: {str(e)}')
        
        # Calculate percentages for all recipes after upload
        if success_count > 0:
            db.session.flush()  # Ensure all recipes are in the database
            
            # Get all unique recipe codes that were uploaded
            uploaded_recipe_codes = set()
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if any(row):
                    row_data = dict(zip(headers, row))
                    if row_data.get('Recipe Code'):
                        uploaded_recipe_codes.add(str(row_data['Recipe Code']).strip())
            
            # Recalculate percentages for each recipe group
            for recipe_code in uploaded_recipe_codes:
                recipes_in_group = RecipeMaster.query.filter_by(recipe_code=recipe_code).all()
                total_kg = sum(float(r.kg_per_batch) for r in recipes_in_group)
                
                if total_kg > 0:
                    for recipe in recipes_in_group:
                        recipe.percentage = round((float(recipe.kg_per_batch) / total_kg) * 100, 2)
        
        db.session.commit()
        
        message = f'Upload completed: {success_count} recipes added successfully'
        if error_count > 0:
            message += f', {error_count} errors occurred'
            if len(errors) <= 5:  # Show first 5 errors
                message += f'. Errors: {"; ".join(errors)}'
        
        return jsonify({'message': message}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 500

@recipe_bp.route('/recipe/download-excel', methods=['GET'])
def download_recipe_excel():
    try:
        # Get search parameters
        search_recipe_code = request.args.get('recipe_code', '').strip()
        search_description = request.args.get('description', '').strip()
        
        # Build query with same logic as get_search_recipes
        from sqlalchemy.orm import aliased
        RawMaterialItem = aliased(ItemMaster)
        FinishedGoodItem = aliased(ItemMaster)
        
        query = db.session.query(
            RecipeMaster,
            RawMaterialItem.item_code.label('raw_material_code'),
            RawMaterialItem.description.label('raw_material_name'),
            FinishedGoodItem.item_code.label('finished_good_code'),
            FinishedGoodItem.description.label('finished_good_name')
        ).join(
            RawMaterialItem,
            RecipeMaster.raw_material_id == RawMaterialItem.id
        ).join(
            FinishedGoodItem,
            RecipeMaster.finished_good_id == FinishedGoodItem.id
        )
        
        if search_recipe_code:
            query = query.filter(RecipeMaster.recipe_code.ilike(f"%{search_recipe_code}%"))
        if search_description:
            query = query.filter(RecipeMaster.description.ilike(f"%{search_description}%"))
        
        recipes = query.all()
        
        # Create workbook
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Recipe Master"
        
        # Define headers
        headers = [
            'Recipe Code', 'Description', 'Finished Good Code', 'Finished Good Name',
            'Raw Material Code', 'Raw Material Name', 'Kg Per Batch', 'Percentage', 
            'UOM'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, recipe_data in enumerate(recipes, 2):
            recipe = recipe_data.RecipeMaster
            sheet.cell(row=row, column=1, value=recipe.recipe_code)
            sheet.cell(row=row, column=2, value=recipe.description)
            sheet.cell(row=row, column=3, value=recipe_data.finished_good_code)
            sheet.cell(row=row, column=4, value=recipe_data.finished_good_name)
            sheet.cell(row=row, column=5, value=recipe_data.raw_material_code)
            sheet.cell(row=row, column=6, value=recipe_data.raw_material_name)
            sheet.cell(row=row, column=7, value=recipe.kg_per_batch)
            sheet.cell(row=row, column=8, value=recipe.percentage)
            
            # UOM
            uom_name = ''
            if recipe.quantity_uom_id:
                from models.uom import UOM
                uom = UOM.query.get(recipe.quantity_uom_id)
                if uom:
                    uom_name = uom.UOMName
            sheet.cell(row=row, column=9, value=uom_name)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        import io
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f'recipe_master_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate Excel file: {str(e)}'}), 500

@recipe_bp.route('/recipe/download-template', methods=['GET'])
def download_recipe_template():
    try:
        # Create workbook with template structure
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Recipe Master Template"
        
        # Define headers
        headers = [
            'Recipe Code', 'Description', 'Finished Good Code', 'Raw Material Code',
            'Kg Per Batch', 'UOM'
        ]
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data row
        sample_data = [
            'REC001', 'Sample Recipe Description', 'FG001', 'RM001', '10.5', 'KG'
        ]
        
        for col, value in enumerate(sample_data, 1):
            sheet.cell(row=2, column=col, value=value)
        
        # Add instructions in a separate sheet
        instructions_sheet = workbook.create_sheet("Instructions")
        instructions = [
            "RECIPE MASTER UPLOAD INSTRUCTIONS",
            "",
            "Required Columns:",
            "- Recipe Code: Unique identifier for the recipe group",
            "- Description: Recipe description",
            "- Finished Good Code: Item code of what is being made (must exist in Item Master)",
            "- Raw Material Code: Item code of component needed (must exist in Item Master)",
            "- Kg Per Batch: Quantity of component needed per batch (must be > 0)",
            "",
            "Optional Columns:",
            "- UOM: Unit of measure (must match existing UOMs)",
            "- Is Active: Yes/No (defaults to Yes if not specified)",
            "",
            "Notes:",
            "- Do not modify the header row",
            "- Empty rows will be skipped",
            "- Item codes must exist in the Item Master",
            "- Duplicate finished good + raw material combinations will be rejected",
            "- Percentages will be calculated automatically within each recipe group",
            "- Boolean fields accept: Yes/No, True/False, 1/0, Y/N",
            "",
            "Example:",
            "Recipe Code: FRANK001",
            "Description: Frankfurter Recipe",
            "Finished Good Code: 2006 (Frankfurter WIP)",
            "Raw Material Code: RM001 (Pork 80CL)",
            "Kg Per Batch: 40.5"
        ]
        
        for row, instruction in enumerate(instructions, 1):
            cell = instructions_sheet.cell(row=row, column=1, value=instruction)
            if row == 1:  # Title
                cell.font = Font(bold=True, size=14)
            elif instruction.endswith(":"):  # Section headers
                cell.font = Font(bold=True)
        
        # Auto-adjust column widths for both sheets
        for sheet_obj in [sheet, instructions_sheet]:
            for column in sheet_obj.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                sheet_obj.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        import io
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='recipe_master_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate template: {str(e)}'}), 500

@recipe_bp.route('/recipe/upload', methods=['GET', 'POST'])
def recipe_upload():
    """Render the upload page for Recipe Master Excel upload and handle file uploads."""
    if request.method == 'POST':
        try:
            if 'file' not in request.files:
                flash('No file uploaded', 'error')
                return render_template('recipe/upload.html', current_page='recipe')
            
            file = request.files['file']
            if file.filename == '':
                flash('No file selected', 'error')
                return render_template('recipe/upload.html', current_page='recipe')
            
            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                flash('Please upload a valid Excel file (.xlsx or .xls)', 'error')
                return render_template('recipe/upload.html', current_page='recipe')
            
            # Get sheet name from form
            sheet_name = request.form.get('sheet_name', '').strip() or 'Recipe Master'
            
            # Process the file using existing upload logic
            import openpyxl
            workbook = openpyxl.load_workbook(file)
            
            # Try to get the specified sheet
            try:
                sheet = workbook[sheet_name]
            except KeyError:
                # If specified sheet doesn't exist, try active sheet
                sheet = workbook.active
                flash(f'Sheet "{sheet_name}" not found, using active sheet instead', 'warning')
            
            # Expected columns
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # Validate required headers
            required_headers = ['Recipe Code', 'Description', 'Finished Good Code', 'Raw Material Code', 'Kg Per Batch']
            missing_headers = [h for h in required_headers if h not in headers]
            if missing_headers:
                flash(f'Missing required columns: {", ".join(missing_headers)}', 'error')
                return render_template('recipe/upload.html', current_page='recipe')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Process each row (skip header)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # Skip empty rows
                        continue
                    
                    # Create dictionary from row data
                    row_data = dict(zip(headers, row))
                    
                    # Skip if required fields are empty
                    if not all([row_data.get('Recipe Code'), row_data.get('Description'), 
                               row_data.get('Finished Good Code'), row_data.get('Raw Material Code'), 
                               row_data.get('Kg Per Batch')]):
                        error_count += 1
                        errors.append(f'Row {row_num}: Missing required fields')
                        continue
                    
                    recipe_code = str(row_data['Recipe Code']).strip()
                    description = str(row_data['Description']).strip()
                    finished_good_code = str(row_data['Finished Good Code']).strip()
                    raw_material_code = str(row_data['Raw Material Code']).strip()
                    
                    # Validate kg_per_batch is a number
                    try:
                        kg_per_batch = float(row_data['Kg Per Batch'])
                        if kg_per_batch <= 0:
                            error_count += 1
                            errors.append(f'Row {row_num}: Kg per batch must be greater than 0')
                            continue
                    except (ValueError, TypeError):
                        error_count += 1
                        errors.append(f'Row {row_num}: Invalid kg per batch value')
                        continue
                    
                    # Find finished good item
                    finished_good = ItemMaster.query.filter_by(item_code=finished_good_code).first()
                    if not finished_good:
                        error_count += 1
                        errors.append(f'Row {row_num}: Finished good code "{finished_good_code}" not found')
                        continue
                    
                    # Find raw material item
                    raw_material = ItemMaster.query.filter_by(item_code=raw_material_code).first()
                    if not raw_material:
                        error_count += 1
                        errors.append(f'Row {row_num}: Raw material code "{raw_material_code}" not found')
                        continue
                    
                    # Check if recipe already exists
                    existing_recipe = RecipeMaster.query.filter(
                        RecipeMaster.finished_good_id == finished_good.id,
                        RecipeMaster.raw_material_id == raw_material.id
                    ).first()
                    if existing_recipe:
                        error_count += 1
                        errors.append(f'Row {row_num}: Recipe for {finished_good_code} -> {raw_material_code} already exists')
                        continue
                    
                    # Find UOM if specified
                    uom_id = None
                    if row_data.get('UOM'):
                        from models.uom import UOM
                        uom = UOM.query.filter_by(UOMName=str(row_data['UOM']).strip()).first()
                        if uom:
                            uom_id = uom.UOMID
                    
                    # Create new recipe
                    recipe = RecipeMaster()
                    recipe.recipe_code = recipe_code
                    recipe.description = description
                    recipe.finished_good_id = finished_good.id
                    recipe.raw_material_id = raw_material.id
                    recipe.kg_per_batch = kg_per_batch
                    recipe.quantity_uom_id = uom_id
                    
                    db.session.add(recipe)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Row {row_num}: {str(e)}')
            
            # Calculate percentages for all recipes after upload
            if success_count > 0:
                db.session.flush()  # Ensure all recipes are in the database
                
                # Get all unique recipe codes that were uploaded
                uploaded_recipe_codes = set()
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if any(row):
                        row_data = dict(zip(headers, row))
                        if row_data.get('Recipe Code'):
                            uploaded_recipe_codes.add(str(row_data['Recipe Code']).strip())
                
                # Recalculate percentages for each recipe group
                for recipe_code in uploaded_recipe_codes:
                    recipes_in_group = RecipeMaster.query.filter_by(recipe_code=recipe_code).all()
                    total_kg = sum(float(r.kg_per_batch) for r in recipes_in_group)
                    
                    if total_kg > 0:
                        for recipe in recipes_in_group:
                            recipe.percentage = round((float(recipe.kg_per_batch) / total_kg) * 100, 2)
            
            db.session.commit()
            
            message = f'Upload completed: {success_count} recipes added successfully'
            if error_count > 0:
                message += f', {error_count} errors occurred'
                if len(errors) <= 5:  # Show first 5 errors
                    message += f'. Errors: {"; ".join(errors)}'
                flash(message, 'warning')
            else:
                flash(message, 'success')
            
            return render_template('recipe/upload.html', current_page='recipe')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Failed to process file: {str(e)}', 'error')
            return render_template('recipe/upload.html', current_page='recipe')
    
    # GET request - just render the upload page
    return render_template('recipe/upload.html', current_page='recipe')

@recipe_bp.route('/raw_material_report', methods=['GET'])
def raw_material_report():
    try:
        # Get week commencing filter from request
        week_commencing = request.args.get('week_commencing')
        
        # If no week filter provided, get data from raw_material_report table
        if not week_commencing:
            # Get existing data from raw_material_report table
            raw_material_reports = RawMaterialReport.query.order_by(RawMaterialReport.week_commencing.desc()).all()
            
            raw_material_data = [
                {
                    'week_commencing': report.week_commencing.strftime('%d/%m/%Y'),
                    'raw_material': report.raw_material,
                    'usage': round(float(report.meat_required), 2)
                }
                for report in raw_material_reports
            ]
        else:
            # Week filter provided - recalculate and save new data
            # Base query for weekly data - using current schema with corrected field names
            raw_material_query = """
            SELECT 
                DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY) as week_commencing,
                component_im.description as component_material,
                component_im.id as component_item_id,
                SUM(p.total_kg * (r.quantity_kg / recipe_totals.total_recipe_kg) * 100) as total_usage
            FROM production p
            JOIN item_master production_im ON p.item_id = production_im.id
            JOIN recipe_master r ON production_im.id = r.recipe_wip_id
            JOIN item_master component_im ON r.component_item_id = component_im.id
            JOIN (
                SELECT 
                    r2.recipe_wip_id,
                    SUM(r2.quantity_kg) as total_recipe_kg
                FROM recipe_master r2
                GROUP BY r2.recipe_wip_id
            ) recipe_totals ON r.recipe_wip_id = recipe_totals.recipe_wip_id
            WHERE DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY) = :week_commencing
            GROUP BY 
                DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY),
                component_im.description,
                component_im.id
            ORDER BY week_commencing DESC, component_im.description
            """
            
            params = {'week_commencing': datetime.strptime(week_commencing, '%Y-%m-%d').date()}
            results = db.session.execute(text(raw_material_query), params).fetchall()
            
            # Clear existing records for the week
            delete_query = "DELETE FROM raw_material_report_table WHERE week_commencing = :week_commencing"
            delete_params = {'week_commencing': datetime.strptime(week_commencing, '%Y-%m-%d').date()}
            db.session.execute(text(delete_query), delete_params)
            
            # Save results to raw_material_report table
            for result in results:
                report = RawMaterialReport(
                    production_date=result.week_commencing,  # Using week_commencing as production_date
                    week_commencing=result.week_commencing,
                    raw_material_id=result.component_item_id,  # Matches database structure
                    raw_material=result.component_material,
                    meat_required=float(result.total_usage),  # Matches database structure
                    created_at=datetime.now()
                )
                db.session.add(report)
            
            db.session.commit()
            
            # Convert to list of dictionaries for template
            raw_material_data = [
                {
                    'week_commencing': result.week_commencing.strftime('%d/%m/%Y'),
                    'raw_material': result.component_material,
                    'usage': round(float(result.total_usage), 2)
                }
                for result in results
            ]
        
        return render_template('recipe/raw_material_report.html', 
                             raw_material_data=raw_material_data,
                             week_commencing=week_commencing,
                             current_page='raw_material_report')
        
    except Exception as e:
        flash(f"An error occurred: {str(e)}", 'error')
        return render_template('recipe/raw_material_report.html', 
                             raw_material_data=[],
                             week_commencing=week_commencing,
                             current_page='raw_material_report')

@recipe_bp.route('/raw_material_download', methods=['GET'])
def raw_material_download():
    try:
        week_commencing = request.args.get('week_commencing')
        
        # Base query for weekly data - using current schema with corrected field names
        raw_material_query = """
        SELECT 
            DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY) as week_commencing,
            component_im.description as component_material,
            component_im.id as component_item_id,
            SUM(p.total_kg * (r.quantity_kg / recipe_totals.total_recipe_kg) * 100) as total_usage
        FROM production p
        JOIN item_master production_im ON p.item_id = production_im.id
        JOIN recipe_master r ON production_im.id = r.recipe_wip_id
        JOIN item_master component_im ON r.component_item_id = component_im.id
        JOIN (
            SELECT 
                r2.recipe_wip_id,
                SUM(r2.quantity_kg) as total_recipe_kg
            FROM recipe_master r2
            GROUP BY r2.recipe_wip_id
        ) recipe_totals ON r.recipe_wip_id = recipe_totals.recipe_wip_id
        """
        
        # Add date filter to the query
        params = {}
        if week_commencing:
            raw_material_query += """ 
            WHERE DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY) = :week_commencing
            """
            params['week_commencing'] = datetime.strptime(week_commencing, '%Y-%m-%d').date()
        
        raw_material_query += """
        GROUP BY 
            DATE(p.production_date - INTERVAL (WEEKDAY(p.production_date)) DAY),
            component_im.description,
            component_im.id
        ORDER BY week_commencing DESC, component_im.description
        """
        
        results = db.session.execute(text(raw_material_query), params).fetchall()
        
        # Convert to list of dictionaries for Excel
        data = [
            {
                'Week Commencing': result.week_commencing.strftime('%d/%m/%Y'),
                'Component Material': result.component_material,
                'Total Usage (kg)': round(float(result.total_usage), 2)
            }
            for result in results
        ]
        
        if not data:
            flash("No data available for the selected week.", 'warning')
            return redirect(url_for('recipe.raw_material_report'))
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Raw Material Report', index=False)
            
            # Get the workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Raw Material Report']
            
            # Add formatting
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#D7E4BC',
                'border': 1
            })
            
            # Write the column headers with the defined format
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Adjust column widths
            worksheet.set_column('A:A', 15)  # Week Commencing
            worksheet.set_column('B:B', 30)  # Component Material
            worksheet.set_column('C:C', 15)  # Total Usage
        
        output.seek(0)
        
        filename = f'raw_material_report_{week_commencing}.xlsx' if week_commencing else 'raw_material_report.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f"Error generating Excel file: {str(e)}", 'error')
        return redirect(url_for('recipe.raw_material_report')) 

@recipe_bp.route('/autocomplete_items', methods=['GET'])
def autocomplete_items():
    """Autocomplete for ItemMaster items in recipes"""
    search = request.args.get('query', '').strip()
    item_type = request.args.get('type', '').strip()  # Optional filter by item type
    
    if not search:
        return jsonify([])
    
    try:
        # Base query
        query = ItemMaster.query.filter(
            ItemMaster.item_code.ilike(f"%{search}%") | 
            ItemMaster.description.ilike(f"%{search}%")
        )
        
        # Optional filter by item type
        if item_type:
            from models.item_type import ItemType
            item_type_obj = ItemType.query.filter_by(type_name=item_type).first()
            if item_type_obj:
                query = query.filter(ItemMaster.item_type_id == item_type_obj.id)
        
        # Limit results
        items = query.limit(25).all()
        
        suggestions = [
            {
                "id": item.id,
                "item_code": item.item_code,
                "description": item.description,
                "item_type": item.item_type.type_name if item.item_type else None,
                "display_text": f"{item.item_code} - {item.description}"
            }
            for item in items
        ]
        
        return jsonify(suggestions)
    except Exception as e:
        print(f"Error fetching item autocomplete suggestions: {e}")
        return jsonify([])

@recipe_bp.route('/get_item_by_id/<int:item_id>', methods=['GET'])
def get_item_by_id(item_id):
    """Get specific item details by ID"""
    try:
        item = ItemMaster.query.get_or_404(item_id)
        return jsonify({
            "id": item.id,
            "item_code": item.item_code,
            "description": item.description,
            "item_type": item.item_type.type_name if item.item_type else None,
            "display_text": f"{item.item_code} - {item.description}"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 404 